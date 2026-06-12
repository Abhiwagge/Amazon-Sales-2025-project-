import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

df = pd.read_csv('/home/claude/amazon_sales_project/data/amazon_sales_2025.csv')

wb = Workbook()

HEADER_FILL = PatternFill('solid', start_color='232F3E')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
TITLE_FONT = Font(bold=True, size=14, name='Arial', color='232F3E')
LABEL_FONT = Font(bold=True, name='Arial', size=11)
NORMAL_FONT = Font(name='Arial', size=10)
ACCENT_FILL = PatternFill('solid', start_color='FF9900')
thin = Side(border_style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ============================================================
# SHEET 1: RAW DATA
# ============================================================
ws_raw = wb.active
ws_raw.title = 'Raw_Data'

for c_idx, col in enumerate(df.columns, 1):
    cell = ws_raw.cell(row=1, column=c_idx, value=col)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

for r_idx, row in enumerate(df.itertuples(index=False), 2):
    for c_idx, val in enumerate(row, 1):
        ws_raw.cell(row=r_idx, column=c_idx, value=val)

for c_idx, col in enumerate(df.columns, 1):
    max_len = max(len(str(col)), df[col].astype(str).str.len().max())
    ws_raw.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len + 2, 10), 22)

ws_raw.freeze_panes = 'A2'
n_rows = len(df) + 1

# Define table reference range as a name for formulas
RAW_RANGE = f"Raw_Data!$A$2:$T${n_rows}"

# Column letter map
cols = list(df.columns)
def col_letter(name):
    return get_column_letter(cols.index(name) + 1)

NET_COL = col_letter('Net_Amount')
CAT_COL = col_letter('Category')
STATUS_COL = col_letter('Order_Status')
SEG_COL = col_letter('Customer_Segment')
CITY_COL = col_letter('City')
PROD_COL = col_letter('Product')
QTY_COL = col_letter('Quantity')
DISC_COL = col_letter('Discount_Pct')
RATING_COL = col_letter('Rating')
DATE_COL = col_letter('Order_Date')
PAY_COL = col_letter('Payment_Method')
CHANNEL_COL = col_letter('Sales_Channel')

# ============================================================
# SHEET 2: DASHBOARD
# ============================================================
ws_dash = wb.create_sheet('Dashboard')
ws_dash.sheet_view.showGridLines = False

ws_dash['B2'] = 'AMAZON SALES 2025 — EXECUTIVE DASHBOARD'
ws_dash['B2'].font = Font(bold=True, size=16, color='232F3E', name='Arial')
ws_dash.merge_cells('B2:H2')

kpi_labels = [
    ('Total Net Revenue (₹)', f"=SUMIF(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},\"<>Cancelled\",Raw_Data!${NET_COL}$2:${NET_COL}${n_rows})", '#,##0'),
    ('Total Orders (Valid)', f"=COUNTIF(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},\"<>Cancelled\")", '#,##0'),
    ('Average Order Value (₹)', None, '#,##0.00'),
    ('Return Rate (%)', f"=COUNTIF(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},\"Returned\")/COUNTA(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows})", '0.00%'),
    ('Cancellation Rate (%)', f"=COUNTIF(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},\"Cancelled\")/COUNTA(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows})", '0.00%'),
    ('Total Units Sold', f"=SUMIF(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},\"<>Cancelled\",Raw_Data!${QTY_COL}$2:${QTY_COL}${n_rows})", '#,##0'),
    ('Avg Discount Given (%)', f"=AVERAGE(Raw_Data!${DISC_COL}$2:${DISC_COL}${n_rows})/100", '0.0%'),
    ('Avg Customer Rating', f"=AVERAGE(Raw_Data!${RATING_COL}$2:${RATING_COL}${n_rows})", '0.00'),
]

row = 4
for i, (label, formula, fmt) in enumerate(kpi_labels):
    col_offset = (i % 4) * 2
    r = row + (i // 4) * 3
    c = 2 + col_offset
    lbl_cell = ws_dash.cell(row=r, column=c, value=label)
    lbl_cell.font = Font(bold=True, size=9, color='666666', name='Arial')
    val_cell = ws_dash.cell(row=r+1, column=c, value=formula)
    val_cell.font = Font(bold=True, size=14, color='FF9900', name='Arial')
    val_cell.number_format = fmt
    ws_dash.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
    ws_dash.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+1)

# AOV = Total Revenue / Total Orders (reference the two cells just written)
ws_dash['D5'] = '=B5/D5_placeholder'  # placeholder, fix below
# Fix AOV formula referencing actual cells (B5 = revenue, D5 = orders)
ws_dash['D5'] = '=B5/D5'  # will be overwritten properly below

# Recompute positions properly: kpi_labels order -> positions
# i=0 -> r=4,c=2 (Revenue) -> value cell B5
# i=1 -> r=4,c=4 (Orders)  -> value cell D5
# i=2 -> r=4,c=6 (AOV)     -> value cell F5  = B5 / D5
# i=3 -> r=4,c=8 (Return)  -> value cell H5
# i=4 -> r=7,c=2 (Cancel)  -> value cell B8
# i=5 -> r=7,c=4 (Units)   -> value cell D8
# i=6 -> r=7,c=6 (Disc)    -> value cell F8
# i=7 -> r=7,c=8 (Rating)  -> value cell H8
ws_dash['F5'] = '=B5/D5'
ws_dash['F5'].number_format = '#,##0.00'
ws_dash['F5'].font = Font(bold=True, size=14, color='FF9900', name='Arial')

for r in [4,7]:
    for c in [2,4,6,8]:
        ws_dash.cell(row=r, column=c).fill = PatternFill('solid', start_color='F3F3F3')
        ws_dash.cell(row=r+1, column=c).fill = PatternFill('solid', start_color='F3F3F3')

# ============================================================
# SHEET 3: PIVOT - CATEGORY SUMMARY
# ============================================================
ws_cat = wb.create_sheet('Pivot_Category')
ws_cat['A1'] = 'Category-wise Revenue Summary'
ws_cat['A1'].font = TITLE_FONT

categories = sorted(df['Category'].unique())
headers = ['Category', 'Orders', 'Units Sold', 'Net Revenue (₹)', '% of Total Revenue', 'Avg Order Value']
for c_idx, h in enumerate(headers, 1):
    cell = ws_cat.cell(row=3, column=c_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for i, cat in enumerate(categories):
    r = 4 + i
    ws_cat.cell(row=r, column=1, value=cat)
    ws_cat.cell(row=r, column=2, value=f'=COUNTIFS(Raw_Data!${CAT_COL}$2:${CAT_COL}${n_rows},A{r},Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},"<>Cancelled")')
    ws_cat.cell(row=r, column=3, value=f'=SUMIFS(Raw_Data!${QTY_COL}$2:${QTY_COL}${n_rows},Raw_Data!${CAT_COL}$2:${CAT_COL}${n_rows},A{r},Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},"<>Cancelled")')
    ws_cat.cell(row=r, column=4, value=f'=SUMIFS(Raw_Data!${NET_COL}$2:${NET_COL}${n_rows},Raw_Data!${CAT_COL}$2:${CAT_COL}${n_rows},A{r},Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},"<>Cancelled")')
    ws_cat.cell(row=r, column=4).number_format = '#,##0'
    ws_cat.cell(row=r, column=5, value=f'=D{r}/SUM($D$4:$D${3+len(categories)})')
    ws_cat.cell(row=r, column=5).number_format = '0.0%'
    ws_cat.cell(row=r, column=6, value=f'=D{r}/B{r}')
    ws_cat.cell(row=r, column=6).number_format = '#,##0.00'

last_cat_row = 3 + len(categories)
ws_cat.cell(row=last_cat_row+1, column=1, value='TOTAL').font = LABEL_FONT
for col_letter_, c_idx in zip(['B','C','D'], [2,3,4]):
    ws_cat.cell(row=last_cat_row+1, column=c_idx, value=f'=SUM({col_letter_}4:{col_letter_}{last_cat_row})')
    ws_cat.cell(row=last_cat_row+1, column=c_idx).font = LABEL_FONT
    if c_idx == 4:
        ws_cat.cell(row=last_cat_row+1, column=c_idx).number_format = '#,##0'

for c_idx in range(1,7):
    ws_cat.column_dimensions[get_column_letter(c_idx)].width = 20

# Bar chart
chart1 = BarChart()
chart1.title = "Revenue by Category"
chart1.style = 10
data = Reference(ws_cat, min_col=4, min_row=3, max_row=last_cat_row)
cats = Reference(ws_cat, min_col=1, min_row=4, max_row=last_cat_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.height = 9
chart1.width = 16
ws_cat.add_chart(chart1, 'H3')

# ============================================================
# SHEET 4: PIVOT - MONTHLY TREND
# ============================================================
ws_month = wb.create_sheet('Pivot_Monthly')
ws_month['A1'] = 'Monthly Revenue Trend'
ws_month['A1'].font = TITLE_FONT

month_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
headers2 = ['Month', 'Orders', 'Net Revenue (₹)', 'Avg Order Value']
for c_idx, h in enumerate(headers2, 1):
    cell = ws_month.cell(row=3, column=c_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

for i, m in enumerate(month_names, 1):
    r = 3 + i
    ws_month.cell(row=r, column=1, value=m)
    # MONTH(date) match
    ws_month.cell(row=r, column=2, value=(
        f'=SUMPRODUCT((MONTH(Raw_Data!${DATE_COL}$2:${DATE_COL}${n_rows})={i})*'
        f'(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows}<>"Cancelled"))'
    ))
    ws_month.cell(row=r, column=3, value=(
        f'=SUMPRODUCT((MONTH(Raw_Data!${DATE_COL}$2:${DATE_COL}${n_rows})={i})*'
        f'(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows}<>"Cancelled")*'
        f'Raw_Data!${NET_COL}$2:${NET_COL}${n_rows})'
    ))
    ws_month.cell(row=r, column=3).number_format = '#,##0'
    ws_month.cell(row=r, column=4, value=f'=C{r}/B{r}')
    ws_month.cell(row=r, column=4).number_format = '#,##0.00'

for c_idx in range(1,5):
    ws_month.column_dimensions[get_column_letter(c_idx)].width = 16

line1 = LineChart()
line1.title = "Monthly Net Revenue - 2025"
data = Reference(ws_month, min_col=3, min_row=3, max_row=15)
cats = Reference(ws_month, min_col=1, min_row=4, max_row=15)
line1.add_data(data, titles_from_data=True)
line1.set_categories(cats)
line1.height = 9
line1.width = 16
ws_month.add_chart(line1, 'F3')

# ============================================================
# SHEET 5: PIVOT - CITY / SEGMENT / CHANNEL
# ============================================================
ws_other = wb.create_sheet('Pivot_City_Segment')
ws_other['A1'] = 'City-wise Revenue (Top 10)'
ws_other['A1'].font = TITLE_FONT

city_rev = df[df['Order_Status']!='Cancelled'].groupby('City')['Net_Amount'].sum().sort_values(ascending=False).head(10)
headers3 = ['City', 'Net Revenue (₹)']
for c_idx, h in enumerate(headers3, 1):
    cell = ws_other.cell(row=3, column=c_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

for i, city in enumerate(city_rev.index):
    r = 4 + i
    ws_other.cell(row=r, column=1, value=city)
    ws_other.cell(row=r, column=2, value=f'=SUMIFS(Raw_Data!${NET_COL}$2:${NET_COL}${n_rows},Raw_Data!${CITY_COL}$2:${CITY_COL}${n_rows},A{r},Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},"<>Cancelled")')
    ws_other.cell(row=r, column=2).number_format = '#,##0'

ws_other['D1'] = 'Customer Segment Summary'
ws_other['D1'].font = TITLE_FONT
headers4 = ['Segment', 'Orders', 'Net Revenue (₹)', 'AOV']
for c_idx, h in enumerate(headers4, 4):
    cell = ws_other.cell(row=3, column=c_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

for i, seg in enumerate(['Prime', 'Non-Prime']):
    r = 4 + i
    ws_other.cell(row=r, column=4, value=seg)
    ws_other.cell(row=r, column=5, value=f'=COUNTIFS(Raw_Data!${SEG_COL}$2:${SEG_COL}${n_rows},D{r},Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},"<>Cancelled")')
    ws_other.cell(row=r, column=6, value=f'=SUMIFS(Raw_Data!${NET_COL}$2:${NET_COL}${n_rows},Raw_Data!${SEG_COL}$2:${SEG_COL}${n_rows},D{r},Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},"<>Cancelled")')
    ws_other.cell(row=r, column=6).number_format = '#,##0'
    ws_other.cell(row=r, column=7, value=f'=F{r}/E{r}')
    ws_other.cell(row=r, column=7).number_format = '#,##0.00'

ws_other['I1'] = 'Sales Channel Summary'
ws_other['I1'].font = TITLE_FONT
headers5 = ['Channel', 'Orders', 'Net Revenue (₹)']
for c_idx, h in enumerate(headers5, 9):
    cell = ws_other.cell(row=3, column=c_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

for i, ch in enumerate(['Amazon App','Amazon Web','Alexa Voice Order']):
    r = 4 + i
    ws_other.cell(row=r, column=9, value=ch)
    ws_other.cell(row=r, column=10, value=f'=COUNTIFS(Raw_Data!${CHANNEL_COL}$2:${CHANNEL_COL}${n_rows},I{r},Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},"<>Cancelled")')
    ws_other.cell(row=r, column=11, value=f'=SUMIFS(Raw_Data!${NET_COL}$2:${NET_COL}${n_rows},Raw_Data!${CHANNEL_COL}$2:${CHANNEL_COL}${n_rows},I{r},Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},"<>Cancelled")')
    ws_other.cell(row=r, column=11).number_format = '#,##0'

for c_idx in range(1,12):
    ws_other.column_dimensions[get_column_letter(c_idx)].width = 16

# pie chart for segment
pie1 = PieChart()
pie1.title = "Revenue Share: Prime vs Non-Prime"
data = Reference(ws_other, min_col=6, min_row=4, max_row=5)
cats = Reference(ws_other, min_col=4, min_row=4, max_row=5)
pie1.add_data(data)
pie1.set_categories(cats)
pie1.height = 7
pie1.width = 10
ws_other.add_chart(pie1, 'D8')

# ============================================================
# SHEET 6: ORDER STATUS / RETURNS
# ============================================================
ws_status = wb.create_sheet('Order_Status_Analysis')
ws_status['A1'] = 'Order Status Overview'
ws_status['A1'].font = TITLE_FONT

headers6 = ['Order_Status', 'Count', '% of Total']
for c_idx, h in enumerate(headers6, 1):
    cell = ws_status.cell(row=3, column=c_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

statuses = ['Delivered','Returned','Cancelled','Shipped - In Transit']
for i, st in enumerate(statuses):
    r = 4 + i
    ws_status.cell(row=r, column=1, value=st)
    ws_status.cell(row=r, column=2, value=f'=COUNTIF(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},A{r})')
    ws_status.cell(row=r, column=3, value=f'=B{r}/COUNTA(Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows})')
    ws_status.cell(row=r, column=3).number_format = '0.0%'

ws_status['A10'] = 'Return Rate by Category'
ws_status['A10'].font = TITLE_FONT
headers7 = ['Category', 'Total Orders', 'Returns', 'Return Rate %']
for c_idx, h in enumerate(headers7, 1):
    cell = ws_status.cell(row=12, column=c_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

for i, cat in enumerate(categories):
    r = 13 + i
    ws_status.cell(row=r, column=1, value=cat)
    ws_status.cell(row=r, column=2, value=f'=COUNTIF(Raw_Data!${CAT_COL}$2:${CAT_COL}${n_rows},A{r})')
    ws_status.cell(row=r, column=3, value=f'=COUNTIFS(Raw_Data!${CAT_COL}$2:${CAT_COL}${n_rows},A{r},Raw_Data!${STATUS_COL}$2:${STATUS_COL}${n_rows},"Returned")')
    ws_status.cell(row=r, column=4, value=f'=C{r}/B{r}')
    ws_status.cell(row=r, column=4).number_format = '0.0%'

pie2 = PieChart()
pie2.title = "Order Status Distribution"
data = Reference(ws_status, min_col=2, min_row=4, max_row=7)
cats = Reference(ws_status, min_col=1, min_row=4, max_row=7)
pie2.add_data(data)
pie2.set_categories(cats)
pie2.height = 8
pie2.width = 12
ws_status.add_chart(pie2, 'F3')

for c_idx in range(1,5):
    ws_status.column_dimensions[get_column_letter(c_idx)].width = 22

# Reorder sheets: Dashboard first
wb.move_sheet('Dashboard', offset=-(wb.sheetnames.index('Dashboard')))

wb.save('/home/claude/amazon_sales_project/excel/Amazon_Sales_2025_Analysis.xlsx')
print("Saved Excel workbook.")
